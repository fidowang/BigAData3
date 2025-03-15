import json

from datetime import datetime, timedelta
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


def read_json_file(path: str) -> list | dict:
    with open(path, 'r', encoding='UTF-8') as jsonfile:
        json_data = json.load(jsonfile)
    jsonfile.close()
    return json_data


def write_json_file(path: str, json_obj: list | dict):
    with open(path, 'w', encoding='UTF-8') as file:
        json.dump(json_obj, file, ensure_ascii=False, indent=4)
    file.close()


def date_to_xls_num(date=datetime.today().date()) -> int:
    if isinstance(date, str):
        dt1 = datetime.strptime(date, '%Y-%m-%d').date()
        # dt1 = dt1.date()
    else:
        dt1 = date
    return dt1 - datetime(1900, 1, 1).date() + timedelta(days=2)


def date_str(dt: str = datetime.today().date().strftime('%Y-%m-%d')) -> str:
    return dt


class BaseData(object):
    def __init__(self):
        self.title_row_map: Dict[str, int] = {
            '日期': 0,
            '代码': 1,
            '名称': 2,
            '几天几板': 3,
            '涨停类型': 4,
            '原因类别': 5,
            '涨幅': 6,
            '现价': 7,
            '涨停时间': 8,
            '金额': 9,
            '流值': 10,
            '韭研原因': 11,  # 同花顺导出数据中该列为连续涨停，后变更为韭研原因
            '开票涨幅': 12,
            '竞价金额': 13
        }

        self.date_str = date_str()

        self.dateToExcelQuery = date_to_xls_num()

        self.bigA_path = '.\\bigAReview.xlsx'

        self.bigA_back = f'.\\review_backup\\bigAReview_{self.date_str}.xlsx'

        self.redGreenCount: list = []

        self.zt_counts: int = 0

        self.stock_list: List[list] = []

        self.marketIndexInfo: List = []

        self.marketMoodData: list = []

        self.marketDetialInfo: list = []

    def ths_match_jy(self, thsdata: object, jydata: object) -> list:
        """
        匹配同花顺和韭研的数据，输出结果是同花顺数据，用于完成bigAReview
        :param ths_data: 同花顺导入的数据
        :param jy_data: 韭研导入的数据
        :return: 输出同花顺个股数据列表，用于填入bigAReview
        """
        print('开始匹配同花顺和韭研公社涨停股数据')
        compiledStocksData = pd.merge(thsdata, jydata[['code', 'fieldName', 'briefReason']], left_on='代码', right_on='code', how='left')
        compiledStocksData['所属概念'] = compiledStocksData['所属概念'] + '||' + compiledStocksData['briefReason'].apply(lambda x: '+'.join(x) if isinstance(x, list) else str(x))
        # compiledStocksData['涨停原因类别'] = compiledStocksData['涨停原因类别'] + '||' + compiledStocksData['fieldName']
        compiledStocksData.drop(columns='code', axis='columns', inplace=True)
        # compiledStocksData.drop(columns='fieldName', axis='columns', inplace=True)
        compiledStocksData.drop(columns='briefReason', axis='columns', inplace=True)

        tmpResultData = compiledStocksData.values.tolist()
        print('数据匹配完成')
        print(compiledStocksData.to_string())
        print('\n')
        return tmpResultData

    def xls_edit(self) -> None:
        # book_path = self.bigA_path
        print('开始打开bigAReview文件')
        book_path = load_workbook(self.bigA_path)
        # book_path = load_workbook("new.xlsx")
        print('文件正在进行备份')
        book_path.save(self.bigA_back)
        print('备份完成，开始写入数据')

        # self.marketinfo_wrt(book_path)
        self.marketInfoWrite(book_path)
        # self.limitup_wrt(book_path, self.zt_counts, self.stock_list)

        book_path.save(self.bigA_path)
        # book_path.save("new.xlsx")
        print('数据写入完成')

    def limitup_wrt(self, book_review: object, limitup_counts: int, limit_up_stock: list) -> None:

        print('limitup_wrt 开始运行')
        print('涨停板表格数据开始写入')
        sheet_limituplist = book_review['涨停板']
        sheet_limituplist.insert_rows(2, limitup_counts)
        style_date = '日期'
        style_code = '代码'
        style_time = '时间'
        style_million = '百万'
        if '百分比' not in book_review.named_styles:
            print("无\"百分比\"样式，需建立样式")
            style_percent = NamedStyle(name='百分比')
            style_percent.number_format = "0.00%"
        else:
            style_percent = '百分比'

        if '亿元' not in book_review.named_styles:
            print("无\"亿元\"样式，需建立样式")
            style_currency = NamedStyle(name='亿元')
            style_currency.number_format = "0!.00,,\"亿\""
        else:
            style_currency = '亿元'

        for irow in range(0, limitup_counts):
            for icol in range(0, len(limit_up_stock[0])):
                sheet_limituplist.cell(irow + 2, icol + 1).value = limit_up_stock[irow][icol]
                match icol:
                    case 0:  # 对部分列进行格式化
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_date
                    case 1:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_code
                    case 6:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_time
                    case 7:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_time
                    case 9:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_currency
                    case 10:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_million
                    case 11:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_currency
                    case 12:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_percent
                    case 13:
                        sheet_limituplist.cell(irow + 2, icol + 1).style = style_percent
        # '日期'0 代码1 名称2 几天几板3 连续涨停天数4 所属概念5 首次涨停时间6 最终涨停时间7 现价8 金额9 竞价金额10 自由流值11 涨幅12 开盘涨幅13 同花顺涨停原因类别14 韭研异动板块15

        print('涨停板表格数据已写入')

    def marketInfoWrite(self, book_review: object) -> None:
        print('marketInfoWrite 开始运行')
        titleMap: Dict[str, int] = {
            '日期': 1,
            '成交量': 2,
            '上证涨幅': 3,
            '上涨数': 4,
            '平盘数': 5,
            '下跌数': 6,
            '涨停数（不含st）': 7,
            '首板数': 8,
            '最高板位': 9,
            '炸板数': 10,
            '炸板表现': 11,
            '断板数': 12,
            '断板表现': 13,
            '曾跌停数': 14,
            '跌停数': 15,
            '一字跌停数': 16,
            '连续跌停数': 17,
            '天地板数': 18,
            '地天板数': 19,
            '昨日涨停表现': 20,
            '昨日连板表现': 21,
            '昨日炸板表现': 22,
            '市场高标': 23,
            '连板成功率': 24,
            '连板率': 25,
            '封板率': 26,
        }
        dateToWrite = date_to_xls_num()
        marketInfoSheet = book_review['市场']
        styleWeekdate: str = '日期周'
        stylePercent: str = '百分比'
        rowToWrite: int = 0
        for row in marketInfoSheet.iter_rows(min_col=1, max_col=1):
            # print(f'{row[0].value}')
            if row[0].value == dateToWrite:
                exit('检测到今日日期，请检查今日数据是否已写入')
            else:
                rowToWrite = row[0].row + 1
        marketInfoSheet.append(self.marketDetialInfo)
        marketInfoSheet.cell(rowToWrite, titleMap['日期']).style = styleWeekdate
        marketInfoSheet.cell(rowToWrite, titleMap['连板成功率']).value = f'=(G{rowToWrite}-H{rowToWrite})/G{rowToWrite - 1}'
        marketInfoSheet.cell(rowToWrite, titleMap['连板率']).value = f'=(G{rowToWrite}-H{rowToWrite})/G{rowToWrite}'
        marketInfoSheet.cell(rowToWrite, titleMap['封板率']).value = f'=G{rowToWrite}/(J{rowToWrite}+G{rowToWrite})'

        marketInfoSheet.cell(rowToWrite, titleMap['上证涨幅']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['炸板表现']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['断板表现']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['昨日涨停表现']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['昨日连板表现']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['昨日炸板表现']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['连板成功率']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['连板率']).style = stylePercent
        marketInfoSheet.cell(rowToWrite, titleMap['封板率']).style = stylePercent


if __name__ != '__main__':
    pass
